"""상품 엑셀 → `apps/catalog/fixtures/demo.json` 변환.

    pip install openpyxl
    python scripts/xlsx_to_demo.py --xlsx MCM_전체상품데이터.xlsx

엑셀에 있는 값만 채우고 **없는 축은 빈 문자열로 둔다.** 억지로 추측해 넣으면
틀린 분류가 조용히 박히고, 그 상품은 엉뚱한 사람에게 추천된다. 빈 축은 마지막에
요약으로 알려주니 그걸 보고 채우면 된다.

이미 채워둔 값은 덮어쓰지 않는다. 손으로 축을 채운 뒤 엑셀이 갱신돼도 다시 돌릴 수 있다.
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUT = BASE_DIR / "apps" / "catalog" / "fixtures" / "demo.json"
SHEET_NAME = "MCM_전체상품데이터"

STORE = {"id": "s_mcm", "name": "MCM 스토어"}

# 엑셀 카테고리 → Category 코드. 목록에 없으면 빈 값으로 두고 사람이 정한다.
CATEGORY_MAP = {
    "토트백": "tote",
    "쇼퍼백": "tote",
    "백팩": "backpack",
    "크로스백": "crossbody",
    "숄더백": "shoulder",
    "미니백": "crossbody",
    "지갑": "wallet",
    "악세서리": "accessory",
    "액세서리": "accessory",
    "여성의류": "clothing",
    "남성의류": "clothing",
}

# attributes.color 문자열에 이 조각이 있으면 그 축 값으로 본다. 위에서부터 먼저 맞는 것.
COLOR_RULES = (
    ("visetos", "visetos_mix"),
    ("비세토스", "visetos_mix"),
    ("multi", "visetos_mix"),
    ("배색", "visetos_mix"),
    ("black", "black"),
    ("블랙", "black"),
    ("white", "white"),
    ("화이트", "white"),
    ("cognac", "cognac"),
    ("코냑", "cognac"),
    ("navy", "navy"),
    ("네이비", "navy"),
    ("pink", "pink"),
    ("핑크", "pink"),
    ("lotus", "pink"),
    ("red", "red"),
    ("레드", "red"),
    ("orange", "red"),
    ("silver", "metallic"),
    ("gold", "metallic"),
    ("metallic", "metallic"),
    ("메탈릭", "metallic"),
    ("beige", "beige"),
    ("베이지", "beige"),
    ("khaki", "beige"),
    ("moss", "beige"),
    ("cinnamon", "beige"),
    ("taupe", "beige"),
)

# attributes.material 문자열 기준. 캔버스를 레더보다 먼저 본다 —
# "코팅 캔버스에 레더 트리밍"처럼 둘 다 들어간 설명이 흔하다.
MATERIAL_RULES = (
    ("캔버스", "coated_canvas"),
    ("canvas", "coated_canvas"),
    ("나일론", "nylon"),
    ("nylon", "nylon"),
    ("스웨이드", "suede"),
    ("suede", "suede"),
    ("그레인", "grained_leather"),
    ("grain", "grained_leather"),
    ("레더", "smooth_leather"),
    ("가죽", "smooth_leather"),
    ("카프", "smooth_leather"),
    ("leather", "smooth_leather"),
)

# 이름·스토리에서 무늬를 읽는다. 나머지는 사람이 본다.
PATTERN_RULES = (
    ("비세토스", "visetos_monogram"),
    ("visetos", "visetos_monogram"),
    ("모노그램", "visetos_monogram"),
    ("스터드", "studded"),
)

PRICE_BAND_MID = 600_000
PRICE_BAND_HIGH = 1_200_000

AXES = (
    "category",
    "color",
    "material",
    "pattern",
    "silhouette",
    "mood",
    "price_band",
    "use_case",
)
PRESET_KEYS = ("price", "material", "design_intent")


def main() -> int:
    # 윈도우 기본 콘솔(cp949)에서는 한글 요약과 기호가 그대로 깨진다.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="상품 엑셀을 시드 JSON으로 바꾼다.")
    parser.add_argument("--xlsx", required=True, help="상품 엑셀 경로")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="출력 JSON 경로")
    parser.add_argument("--sheet", default=SHEET_NAME)
    parser.add_argument(
        "--fresh",
        action="store_true",
        help="기존 파일을 참고하지 않고 새로 만든다. 상품 구성이 바뀌어 id가 다른 상품을 "
        "가리키게 됐을 때 쓴다(옛 값이 엉뚱한 상품에 남는 것을 막는다).",
    )
    args = parser.parse_args()

    rows = _read_rows(Path(args.xlsx), args.sheet)
    if not rows:
        print("엑셀에서 읽은 행이 없습니다. --sheet 이름을 확인하세요.")
        return 1

    out_path = Path(args.out)
    previous = {} if args.fresh else _load_previous(out_path)
    scenes = _build_scenes(rows, previous)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_comment": "scripts/xlsx_to_demo.py가 만든 파일. 빈 축은 사람이 채운다.",
        "store": STORE,
        "scenes": scenes,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    _report(scenes, out_path)
    return 0


def _read_rows(xlsx_path: Path, sheet: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl이 필요합니다:  pip install openpyxl")
        raise SystemExit(1) from None

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        print(f"시트 '{sheet}'가 없습니다. 있는 시트: {workbook.sheetnames}")
        raise SystemExit(1)

    records = workbook[sheet].iter_rows(values_only=True)
    header = [str(cell).strip() if cell else "" for cell in next(records)]
    return [dict(zip(header, row, strict=False)) for row in records if row and row[0]]


def _load_previous(out_path: Path) -> dict[str, dict]:
    """이미 만들어 둔 상품을 id로 찾아둔다. 손으로 채운 값을 지우지 않기 위해서다."""
    if not out_path.exists():
        return {}
    data = json.loads(out_path.read_text(encoding="utf-8"))
    return {
        product["id"]: product for scene in data.get("scenes", []) for product in scene.get("products", [])
    }


def _build_scenes(rows: list[dict], previous: dict[str, dict]) -> list[dict]:
    grouped: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        scene_no, product_no = _parse_number(row.get("번호"))
        if scene_no is None:
            continue
        grouped[scene_no].append(_build_product(row, scene_no, product_no, previous))

    return [
        {
            "id": f"sc_{scene_no:02d}",
            "no": scene_no,
            "name": _scene_name(rows, scene_no),
            "products": sorted(products, key=lambda item: item["no"]),
        }
        for scene_no, products in sorted(grouped.items())
    ]


def _parse_number(raw) -> tuple[int | None, int]:
    """번호는 "3-2" = 3번 진열대 2번 상품이다."""
    parts = str(raw or "").split("-")
    if len(parts) != 2 or not all(part.strip().isdigit() for part in parts):
        return None, 0
    return int(parts[0]), int(parts[1])


def _scene_name(rows: list[dict], scene_no: int) -> str:
    """전시존 이름은 그 단의 첫 상품 카테고리를 그대로 쓴다. 기획이 정하면 바꾼다."""
    for row in rows:
        if _parse_number(row.get("번호"))[0] == scene_no:
            return str(row.get("카테고리") or f"{scene_no}단")
    return f"{scene_no}단"


def _build_product(row: dict, scene_no: int, product_no: int, previous: dict[str, dict]) -> dict:
    product_id = f"p_{scene_no}{product_no:02d}"
    old = previous.get(product_id, {})
    attributes = _parse_json(row.get("attributes"), {})
    price = _parse_price(row.get("price"))

    product = {
        "id": product_id,
        "no": product_no,
        "name": str(row.get("name") or "").strip(),
        "price": price,
        "external_url": _keep(old, "external_url", _clean_url(row.get("external_url"))),
        "story": str(row.get("story") or "").strip(),
        "images": _keep(old, "images", _clean_images(row.get("images"))),
        "cutout_url": old.get("cutout_url", ""),
        "is_new": old.get("is_new", scene_no == 7),  # 7단이 F/W 신상이다
        "attributes": attributes or old.get("attributes", {}),
        "preset_answers": _keep_presets(old),
        "llm_context": old.get("llm_context", ""),
    }
    product.update(_axes(row, attributes, price, old))
    return product


def _axes(row: dict, attributes: dict, price: int, old: dict) -> dict:
    """규칙으로 알 수 있는 축만 채운다. 나머지는 빈 문자열로 남겨 사람에게 넘긴다."""
    text = f"{row.get('name') or ''} {row.get('story') or ''}"
    return {
        "category": _keep(old, "category", CATEGORY_MAP.get(str(row.get("카테고리") or "").strip(), "")),
        "color": _keep(old, "color", _match(attributes.get("color"), COLOR_RULES)),
        "material": _keep(old, "material", _match(attributes.get("material"), MATERIAL_RULES)),
        "pattern": _keep(old, "pattern", _match(text, PATTERN_RULES)),
        "price_band": _keep(old, "price_band", _price_band(price)),
        # 이름·설명만으로는 판단이 흔들려 규칙을 두지 않는다.
        "silhouette": old.get("silhouette", ""),
        "mood": old.get("mood", ""),
        "use_case": old.get("use_case", ""),
    }


def _match(value, rules: tuple[tuple[str, str], ...]) -> str:
    lowered = str(value or "").lower()
    for needle, code in rules:
        if needle in lowered:
            return code
    return ""


def _price_band(price: int) -> str:
    if not price:
        return ""
    if price < PRICE_BAND_MID:
        return "entry"
    return "mid" if price <= PRICE_BAND_HIGH else "high"


def _parse_price(raw) -> int:
    digits = "".join(char for char in str(raw or "") if char.isdigit())
    return int(digits) if digits else 0


def _parse_json(raw, fallback):
    try:
        return json.loads(raw) if raw else fallback
    except (TypeError, ValueError):
        return fallback


def _clean_url(raw) -> str:
    """엑셀에 '구매' 같은 표기가 들어 있다. 진짜 주소만 받는다."""
    value = str(raw or "").strip()
    return value if value.startswith("http") else ""


def _clean_images(raw) -> list[str]:
    """["1","2","3"] 같은 자리표시자는 버린다. 빈 목록이면 화면에서 바로 티가 난다."""
    values = _parse_json(raw, [])
    if not isinstance(values, list):
        return []
    return [str(item) for item in values if str(item).startswith("http")]


def _keep(old: dict, key: str, fresh):
    """손으로 채운 값이 있으면 그대로 둔다. 엑셀이 비어 있어도 지우지 않는다."""
    return old.get(key) or fresh


def _keep_presets(old: dict) -> dict:
    stored = old.get("preset_answers", {})
    return {key: stored.get(key, "") for key in PRESET_KEYS}


def _report(scenes: list[dict], out_path: Path) -> None:
    products = [product for scene in scenes for product in scene["products"]]
    print(f"{out_path} 생성 — 전시존 {len(scenes)}개 · 상품 {len(products)}개")
    print()
    print("아직 비어 있는 값 (채워야 취향 분석·화면이 정상 동작한다)")
    for axis in AXES:
        missing = [product["id"] for product in products if not product[axis]]
        if missing:
            sample = ", ".join(missing[:6]) + (" …" if len(missing) > 6 else "")
            print(f"  {axis:<12} {len(missing):>3}개   {sample}")
    for field, label in (("images", "이미지"), ("external_url", "구매 링크"), ("llm_context", "챗봇 설명")):
        missing = [product["id"] for product in products if not product[field]]
        if missing:
            print(f"  {label:<12} {len(missing):>3}개")
    empty_presets = [
        product["id"]
        for product in products
        if any(not product["preset_answers"][key] for key in PRESET_KEYS)
    ]
    if empty_presets:
        print(f"  {'프리셋 3종':<12} {len(empty_presets):>3}개")


if __name__ == "__main__":
    sys.exit(main())
